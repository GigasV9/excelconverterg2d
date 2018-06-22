#   (A,B,C =>Columns on TH.xlsx)
#   csv
#   Issue (4C) number_format doesnt work??
#   Hence user has to change Col K n Col N to 'General' manually

'''
Done
4C  if 5C = USD return ('H')
    elif 5C = BTC return ('K')
    elif 5C = ETH return ('N')
5C  if ('D') ="USD" return USD
    elif ('D')="BTCUSD" return BTC
    elif ('D')="ETHUSD" return ETH
6C  if 5C =(BTC or ETH) return ('H')
8C  =('I')
9C  if 8C not empty return USD
7C  if 6C not empty return USD
2C  Convert type ('C') check with ('F') 
    Credit-Deposit(USD)=>DEPOSIT    if Deposit(BTC)=>TRANSFER
    Debit-Withdrawal(USD)=>WITHDRAW if Withdraw(BTC)=>TRANSFER
3C  if 2C = BUY/SELL,3C = GEMINI
1C  ('A')
'''

from openpyxl import *

fileIN = load_workbook(filename = 'th.xlsx')

ws1 = fileIN['Account History'] 
w2 = fileIN.create_sheet(title="Sheet1")
ws2 = fileIN['Sheet1']

# 5C    Sorts (ws1[Col'4'] into ws2[Col'5'])
for i in range(2,ws1.max_row):
    if ws1.cell(row=i,column=4).value == "ETH":
        ws2.cell(row=i,column=5).value = ws1.cell(row=i,column=4).value
        
    elif ws1.cell(row=i,column=4).value == "ETHUSD":
        ws2.cell(row=i,column=5).value = "ETH"
        
    elif ws1.cell(row=i,column=4).value == "BTC":
        ws2.cell(row=i,column=5).value = ws1.cell(row=i,column=4).value
        
    elif ws1.cell(row=i,column=4).value == "BTCUSD":
        ws2.cell(row=i,column=5).value = "BTC"
        
    else:
        ws2.cell(row=i,column=5).value = "USD"

# 6C    Sorts (ws1[Col'8'] into ws2[Col'6'])
for i in range(2,ws1.max_row):
    if ws2.cell(row=i,column=5).value == "ETH":
        x = ws1.cell(row=i,column=8).value
        if x != None:   # skips none-type before passing abs
            ws2.cell(row=i,column=6).value = abs(x)
            
    elif ws2.cell(row=i,column=5).value == "BTC":
        x = ws1.cell(row=i,column=8).value
        if x != None:   # skips none-type before passing abs
            ws2.cell(row=i,column=6).value = abs(x)

# 4C
for i in range(2,ws1.max_row):
    if ws2.cell(row=i,column=5).value == "USD":
        x = ws1.cell(row=i,column=8).value
        if x != None:   # skips none-type before passing abs
             ws2.cell(row=i,column=4).value = abs(x)

    elif ws2.cell(row=i,column=5).value == "ETH":
        x = ws1.cell(row=i,column=14).value
        ws2.cell(row=i,column=4).value = abs(x)
        
    elif ws2.cell(row=i,column=5).value == "BTC":
        x = ws1.cell(row=i,column=11).value
        ws2.cell(row=i,column=4).value = abs(x)
        
             
# 8C
for i in range(2,ws1.max_row):
    if ws1.cell(row=i,column=9).value != None: # skips none-type before passing abs
        ws2.cell(row=i,column=8).value = abs(ws1.cell(row=i,column=9).value)
    
# 7C
for i in range(2,ws1.max_row):
    if  ws2.cell(row=i,column=6).value != None:
        ws2.cell(row=i,column=7).value = "USD"

# 9C
for i in range(2,ws1.max_row):
    if  ws2.cell(row=i,column=8).value != None:
        ws2.cell(row=i,column=9).value = "USD"

# 2C
for i in range(2,ws1.max_row):
    if ws1.cell(row=i,column=3).value == "Buy":
        ws2.cell(row=i,column=2).value = "BUY"
    
    elif ws1.cell(row=i,column=3).value == "Sell":
        ws2.cell(row=i,column=2).value = "SELL"

    elif ws1.cell(row=i,column=3).value =="Credit":
        if ws1.cell(row=i,column=4).value == "USD":
            ws2.cell(row=i,column=2).value = "DEPOSIT"
        elif ws1.cell(row=i,column=4).value == "BTC" or ws1.cell(row=i,column=4).value == "ETH":
            ws2.cell(row=i,column=2).value = "TRANSFER"

    elif ws1.cell(row=i,column=3).value =="Debit":
        if ws1.cell(row=i,column=4).value == "USD":
            ws2.cell(row=i,column=2).value = "WITHDRAW"
        elif ws1.cell(row=i,column=4).value == "BTC" or ws1.cell(row=i,column=4).value == "ETH":
            ws2.cell(row=i,column=2).value = "TRANSFER"
# 3C
for i in range(2,ws1.max_row):
    if ws2.cell(row=i,column=2).value == "BUY" or ws2.cell(row=i,column=2).value =="DEPOSIT" or ws2.cell(row=i,column=2).value =="SELL":
        ws2.cell(row=i,column=3).value = "Gemini"

# 1C
for i in range(2,ws1.max_row):
    ws2.cell(row=i,column=1).value = ws1.cell(row=i,column=1).value
    

ws2['A1'] = "Date"
ws2['B1'] = "Type"
ws2['C1'] = "Exchange"
ws2['D1'] = "Base amount"
ws2['E1'] = "Base currency"
ws2['F1'] = "Quote amount"
ws2['G1'] = "Quote currency"
ws2['H1'] = "Fee"
ws2['I1'] = "Fee currency"
ws2['J1'] = "Costs/Proceeds"
ws2['K1'] = "Costs/Proceeds currency"
ws2['L1'] = "Sync Holdings"
ws2['M1'] = "Sent/Received from"
ws2['N1'] = "Sent to"
ws2['O1'] = "Notes"

#deletes old worksheet
fileIN.remove(ws1)
# saves results into file
fileIN.save(filename = 'test.xlsx')
